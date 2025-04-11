from flask import Flask, request, jsonify, send_file
import openpyxl
from openpyxl.styles import PatternFill
import os
from datetime import datetime
import tempfile
import shutil
import json
import logging
from pathlib import Path
from config.settings import config

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)

# Load configuration based on environment
env = os.environ.get('FLASK_ENV', 'production')
app.config.from_object(config[env])

# Load cell mapping configuration
def load_cell_mapping():
    try:
        with open(app.config['CELL_MAPPING_PATH'], 'r', encoding='utf-8') as f:
            raw_mapping = json.load(f)
            return convert_coordinates_to_tuples(raw_mapping)
    except FileNotFoundError:
        logger.error(f"Configuration file not found: {app.config['CELL_MAPPING_PATH']}")
        raise RuntimeError(f"Configuration file not found: {app.config['CELL_MAPPING_PATH']}")
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in configuration file: {str(e)}")
        raise RuntimeError(f"Invalid JSON in configuration file: {str(e)}")

def convert_coordinates_to_tuples(data):
    """
    Convert coordinate lists to tuples in the mapping dictionary.
    """
    if isinstance(data, dict):
        return {key: convert_coordinates_to_tuples(value) for key, value in data.items()}
    elif isinstance(data, list) and len(data) == 2:
        return tuple(data)
    return data

# Load cell mapping at startup
try:
    CELL_MAPPING = load_cell_mapping()
    logger.info("Cell mapping configuration loaded successfully")
except Exception as e:
    logger.error(f"Failed to load cell mapping: {str(e)}")
    raise

def validate_template():
    """
    Validate that the template file exists and is accessible.
    """
    if not os.path.exists(app.config['TEMPLATE_PATH']):
        logger.error(f"Template file not found: {app.config['TEMPLATE_PATH']}")
        raise RuntimeError(f"Template file not found: {app.config['TEMPLATE_PATH']}")
    
    try:
        workbook = openpyxl.load_workbook(app.config['TEMPLATE_PATH'])
        if app.config['SHEET_NAME'] not in workbook.sheetnames:
            logger.error(f"Template file does not contain sheet: {app.config['SHEET_NAME']}")
            raise RuntimeError(f"Template file does not contain sheet: {app.config['SHEET_NAME']}")
    except Exception as e:
        logger.error(f"Error validating template file: {str(e)}")
        raise RuntimeError(f"Error validating template file: {str(e)}")

# Validate template at startup
try:
    validate_template()
    logger.info("Template file validated successfully")
except Exception as e:
    logger.error(f"Template validation failed: {str(e)}")
    raise

def transform_data_to_cell_values(data, cell_mapping):
    """
    Transform the dictionary of data into a dictionary of coordinates and values.
    """
    cell_values = {}
    def process_nested_dict(nested_data, nested_mapping):
        for key, value in nested_data.items():
            if key in nested_mapping:
                if isinstance(value, dict):
                    process_nested_dict(value, nested_mapping[key])
                else:
                    cell_values[nested_mapping[key]] = value
            elif key == 'status':
                if value in {"complies", "does_not_comply", "not_applicable"}:
                    cell_values[nested_mapping[value]] = "X"

    process_nested_dict(data, cell_mapping)
    return cell_values

def fill_excel_template(cell_values):
    """
    Fill the Excel with the provided values.
    """
    try:
        # Load the Excel file
        workbook = openpyxl.load_workbook(app.config['TEMPLATE_PATH'])
        sheet = workbook[app.config['SHEET_NAME']]
        
        # Fill the specified cells
        for (row, col), value in cell_values.items():
            if row > 0 and col > 0:
                cell = sheet.cell(row=row, column=col)
                cell.value = value
        
        # Create a temporary file
        temp_dir = tempfile.mkdtemp()
        output_filename = f"filled_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(temp_dir, output_filename)
        
        # Save the modified file
        workbook.save(output_path)
        return output_path, temp_dir
    except Exception as e:
        logger.error(f"Error filling Excel template: {str(e)}")
        raise

@app.route('/fill-template', methods=['POST'])
def fill_template():
    temp_dir = None
    try:
        # Get data from JSON
        data = request.get_json()
        
        if not data:
            logger.warning("Received empty JSON data")
            return jsonify({
                'error': 'Invalid JSON format.'
            }), 400
        
        # Convert coordinates from string to tuples
        cell_values = transform_data_to_cell_values(data, CELL_MAPPING)
        
        # Fill the Excel
        output_path, temp_dir = fill_excel_template(cell_values)
        
        # Return the generated file
        return send_file(
            output_path,
            as_attachment=True,
            download_name=os.path.basename(output_path),
            max_age=0
        )
        
    except Exception as e:
        logger.error(f"Error processing request: {str(e)}")
        return jsonify({
            'error': f'Error processing the request: {str(e)}'
        }), 500
    finally:
        # Clean the temporary directory after sending the response
        if temp_dir and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir, ignore_errors=True)
            except Exception as e:
                logger.error(f"Error cleaning temporary directory: {str(e)}")

if __name__ == '__main__':
    # Use waitress for production
    if env == 'production':
        from waitress import serve
        logger.info("Starting production server with waitress")
        serve(app, host='0.0.0.0', port=5000)
    else:
        logger.info("Starting development server")
        app.run(debug=True, port=5000) 